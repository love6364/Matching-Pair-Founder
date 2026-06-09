
import streamlit as st
import pandas as pd
from io import BytesIO
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ──────────────────────────────────────────
# PAGE CONFIG
# ──────────────────────────────────────────

st.set_page_config(
    page_title="Diamond Pair Finder",
    page_icon="💎",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ──────────────────────────────────────────
# CUSTOM CSS
# ──────────────────────────────────────────

st.markdown("""
<style>
    @import url('https://fonts.googleapis.com/css2?family=DM+Sans:wght@300;400;500;600&display=swap');

    html, body, [class*="css"] {
        font-family: 'DM Sans', sans-serif;
    }

    .stApp {
        background: linear-gradient(135deg, #0f0f1a 0%, #1a1a2e 50%, #16213e 100%);
        min-height: 100vh;
    }

    [data-testid="stSidebar"] {
        background: linear-gradient(180deg, #1a1a2e 0%, #16213e 100%);
        border-right: 1px solid rgba(99, 179, 237, 0.2);
    }

    [data-testid="stSidebar"] .stSlider label,
    [data-testid="stSidebar"] h1,
    [data-testid="stSidebar"] h2,
    [data-testid="stSidebar"] h3,
    [data-testid="stSidebar"] p {
        color: #e2e8f0 !important;
    }

    .stSlider [data-baseweb="slider"] div[role="slider"] {
        background-color: #63b3ed !important;
        border-color: #63b3ed !important;
    }

    .main-title {
        font-size: 2.6rem;
        font-weight: 600;
        background: linear-gradient(135deg, #63b3ed, #9f7aea, #f6ad55);
        -webkit-background-clip: text;
        -webkit-text-fill-color: transparent;
        background-clip: text;
        margin-bottom: 0.2rem;
    }

    .main-subtitle {
        color: #94a3b8;
        font-size: 1rem;
        margin-bottom: 1.5rem;
    }

    [data-testid="metric-container"] {
        background: rgba(255,255,255,0.05);
        border: 1px solid rgba(99,179,237,0.25);
        border-radius: 14px;
        padding: 1rem 1.2rem;
        backdrop-filter: blur(8px);
    }

    [data-testid="metric-container"] [data-testid="stMetricLabel"] {
        color: #94a3b8 !important;
        font-size: 0.78rem !important;
        text-transform: uppercase;
        letter-spacing: 0.08em;
    }

    [data-testid="metric-container"] [data-testid="stMetricValue"] {
        color: #e2e8f0 !important;
        font-size: 2rem !important;
        font-weight: 600 !important;
    }

    [data-testid="stFileUploader"] {
        background: rgba(255,255,255,0.04);
        border: 2px dashed rgba(99,179,237,0.35);
        border-radius: 14px;
        padding: 1rem;
    }

    [data-testid="stFileUploader"] label {
        color: #e2e8f0 !important;
    }

    .stDownloadButton > button,
    .stButton > button {
        background: linear-gradient(135deg, #2b6cb0, #4c51bf) !important;
        color: white !important;
        border: none !important;
        border-radius: 10px !important;
        font-weight: 500 !important;
        font-size: 0.95rem !important;
        padding: 0.55rem 1.4rem !important;
        transition: opacity 0.2s;
        box-shadow: 0 4px 14px rgba(99,179,237,0.3) !important;
    }

    .stDownloadButton > button:hover,
    .stButton > button:hover {
        opacity: 0.88 !important;
    }

    [data-testid="stDataFrame"] {
        border-radius: 12px;
        overflow: hidden;
        border: 1px solid rgba(99,179,237,0.2);
    }

    .stSuccess {
        background: rgba(72,187,120,0.15) !important;
        border: 1px solid rgba(72,187,120,0.4) !important;
        color: #9ae6b4 !important;
        border-radius: 10px !important;
    }

    .stError {
        background: rgba(245,101,101,0.15) !important;
        border: 1px solid rgba(245,101,101,0.4) !important;
        color: #feb2b2 !important;
        border-radius: 10px !important;
    }

    .stInfo {
        background: rgba(99,179,237,0.1) !important;
        border: 1px solid rgba(99,179,237,0.3) !important;
        color: #bee3f8 !important;
        border-radius: 10px !important;
    }

    hr {
        border-color: rgba(99,179,237,0.15) !important;
        margin: 1.2rem 0 !important;
    }

    .section-head {
        font-size: 1.1rem;
        font-weight: 500;
        color: #bee3f8;
        margin-bottom: 0.6rem;
        display: flex;
        align-items: center;
        gap: 6px;
    }

    .tol-badge {
        display: inline-block;
        background: rgba(99,179,237,0.18);
        color: #63b3ed;
        border-radius: 6px;
        padding: 1px 8px;
        font-size: 0.8rem;
        font-weight: 500;
    }

    .streamlit-expanderHeader {
        color: #94a3b8 !important;
        font-size: 0.9rem !important;
    }
</style>
""", unsafe_allow_html=True)

# ──────────────────────────────────────────
# HEADER
# ──────────────────────────────────────────

col_logo, col_title = st.columns([1, 11])
with col_logo:
    st.markdown(
        "<div style='font-size:3rem;margin-top:0.3rem'>💎</div>",
        unsafe_allow_html=True,
    )
with col_title:
    st.markdown(
        "<div class='main-title'>Diamond Pair Finder</div>"
        "<div class='main-subtitle'>Upload your diamond inventory and instantly find matching pairs with configurable tolerances</div>",
        unsafe_allow_html=True,
    )

st.divider()

# ──────────────────────────────────────────
# SIDEBAR — TOLERANCE CONTROLS
# ──────────────────────────────────────────

with st.sidebar:
    st.markdown("## ⚙️ Matching Tolerances")
    st.markdown(
        "<p style='color:#94a3b8;font-size:0.85rem;margin-bottom:1rem'>"
        "Adjust how strict or loose the pairing criteria should be.</p>",
        unsafe_allow_html=True,
    )

    st.markdown("---")

    st.markdown("**📐 Proportions**")
    table_tol = st.slider(
        "Table % tolerance",
        min_value=0.5, max_value=10.0, value=2.0, step=0.5,
        help="Maximum allowed difference in Table %"
    )
    depth_tol = st.slider(
        "Depth % tolerance",
        min_value=0.5, max_value=10.0, value=2.0, step=0.5,
        help="Maximum allowed difference in Depth %"
    )

    st.markdown("---")

    st.markdown("**📏 Dimensions (mm)**")
    length_tol = st.slider(
        "Length tolerance (mm)",
        min_value=0.02, max_value=1.0, value=0.10, step=0.02,
        format="%.2f",
        help="Maximum allowed difference in Length"
    )
    width_tol = st.slider(
        "Width tolerance (mm)",
        min_value=0.02, max_value=1.0, value=0.10, step=0.02,
        format="%.2f",
        help="Maximum allowed difference in Width"
    )

    st.markdown("---")

    st.markdown("**Current settings**")
    st.markdown(
        f"<p style='font-size:0.82rem;color:#94a3b8;line-height:1.8'>"
        f"Table&nbsp;±&nbsp;<b style='color:#63b3ed'>{table_tol}</b>%&nbsp;&nbsp;"
        f"Depth&nbsp;±&nbsp;<b style='color:#63b3ed'>{depth_tol}</b>%<br>"
        f"Length&nbsp;±&nbsp;<b style='color:#63b3ed'>{length_tol:.2f}</b> mm&nbsp;&nbsp;"
        f"Width&nbsp;±&nbsp;<b style='color:#63b3ed'>{width_tol:.2f}</b> mm"
        f"</p>",
        unsafe_allow_html=True,
    )

    st.markdown("---")

    with st.expander("ℹ️ Required columns"):
        st.markdown(
            """
            Your Excel file must contain these columns:
            - `Shape`
            - `Color`
            - `Clarity`
            - `Table %`
            - `Depth %`
            - `Length`
            - `Width`
            """
        )

# ──────────────────────────────────────────
# REQUIRED COLUMNS
# ──────────────────────────────────────────

# SHAPE_COL   = "Shape"
# COLOR_COL   = "Color"
# CLARITY_COL = "Clarity"
# TABLE_COL   = "Table %"
# DEPTH_COL   = "Depth %"
# LENGTH_COL  = "Length"
# WIDTH_COL   = "Width"

# REQUIRED_COLS = [
#     SHAPE_COL, COLOR_COL, CLARITY_COL,
#     TABLE_COL, DEPTH_COL, LENGTH_COL, WIDTH_COL,
# ]
def normalize_col(col):
    return str(col).strip().lower()

COLUMN_ALIASES = {
    "shape": ["shape", "shape name", "shp", "diamond shape"],
    "color": ["color", "col", "colour"],
    "clarity": ["clarity", "cla", "clarity grade"],
    "table": ["table %", "table%", "table", "tbl"],
    "depth": ["depth %", "depth%", "depth", "dep"],
    "length": ["length", "len", "l"],
    "width": ["width", "wid", "w"]
}

def auto_detect_column(columns, aliases):

    normalized = {
        normalize_col(c): c
        for c in columns
    }

    for alias in aliases:

        if normalize_col(alias) in normalized:
            return normalized[
                normalize_col(alias)
            ]

    return None
# ──────────────────────────────────────────
# FILE UPLOAD
# ──────────────────────────────────────────

st.markdown(
    "<div class='section-head'>📂 Upload Inventory File</div>",
    unsafe_allow_html=True,
)

uploaded_file = st.file_uploader(
    "Drag and drop your Excel file here, or click Browse",
    type=["xlsx", "xls"],
    label_visibility="collapsed",
)

if uploaded_file is None:
    st.info("👆 Upload an Excel file (.xlsx / .xls) to get started.")
    st.stop()

# ──────────────────────────────────────────
# READ & VALIDATE
# ──────────────────────────────────────────

try:
    df = pd.read_excel(uploaded_file)
    df.columns = df.columns.str.strip()
except Exception as e:
    st.error(f"❌ Could not read the file: {e}")
    st.stop()

# missing_cols = [c for c in REQUIRED_COLS if c not in df.columns]
# if missing_cols:
#     st.error(f"❌ Missing required columns: **{', '.join(missing_cols)}**")
#     st.stop()
# ==========================================
# AUTO DETECT COLUMNS
# ==========================================

available_cols = list(df.columns)

det_shape = auto_detect_column(
    available_cols,
    COLUMN_ALIASES["shape"]
)

det_color = auto_detect_column(
    available_cols,
    COLUMN_ALIASES["color"]
)

det_clarity = auto_detect_column(
    available_cols,
    COLUMN_ALIASES["clarity"]
)

det_table = auto_detect_column(
    available_cols,
    COLUMN_ALIASES["table"]
)

det_depth = auto_detect_column(
    available_cols,
    COLUMN_ALIASES["depth"]
)

det_length = auto_detect_column(
    available_cols,
    COLUMN_ALIASES["length"]
)

det_width = auto_detect_column(
    available_cols,
    COLUMN_ALIASES["width"]
)

# ==========================================
# COLUMN MAPPING UI
# ==========================================

st.sidebar.markdown("---")
st.sidebar.markdown("## 📋 Column Mapping")

SHAPE_COL = st.sidebar.selectbox(
    "Shape Column",
    available_cols,
    index=available_cols.index(det_shape)
    if det_shape in available_cols else 0
)

COLOR_COL = st.sidebar.selectbox(
    "Color Column",
    available_cols,
    index=available_cols.index(det_color)
    if det_color in available_cols else 0
)

CLARITY_COL = st.sidebar.selectbox(
    "Clarity Column",
    available_cols,
    index=available_cols.index(det_clarity)
    if det_clarity in available_cols else 0
)

TABLE_COL = st.sidebar.selectbox(
    "Table Column",
    available_cols,
    index=available_cols.index(det_table)
    if det_table in available_cols else 0
)

DEPTH_COL = st.sidebar.selectbox(
    "Depth Column",
    available_cols,
    index=available_cols.index(det_depth)
    if det_depth in available_cols else 0
)

LENGTH_COL = st.sidebar.selectbox(
    "Length Column",
    available_cols,
    index=available_cols.index(det_length)
    if det_length in available_cols else 0
)

WIDTH_COL = st.sidebar.selectbox(
    "Width Column",
    available_cols,
    index=available_cols.index(det_width)
    if det_width in available_cols else 0
)

selected_cols = [
    SHAPE_COL,
    COLOR_COL,
    CLARITY_COL,
    TABLE_COL,
    DEPTH_COL,
    LENGTH_COL,
    WIDTH_COL
]

if len(selected_cols) != len(set(selected_cols)):
    st.error(
        "❌ Same column selected multiple times."
    )
    st.stop()
with st.expander(f"📄 Preview uploaded data — {len(df)} rows", expanded=False):
    st.dataframe(df, use_container_width=True, height=220)

# ──────────────────────────────────────────
# PAIR MATCHING
# ──────────────────────────────────────────

# df["Pair_ID"] = ""

# used_rows  = set()
# pair_number = 1

# for i in range(len(df)):
#     if i in used_rows:
#         continue
#     row1 = df.iloc[i]

#     for j in range(i + 1, len(df)):
#         if j in used_rows:
#             continue
#         row2 = df.iloc[j]

#         try:
#             shape_match   = str(row1[SHAPE_COL]).strip()   == str(row2[SHAPE_COL]).strip()
#             color_match   = str(row1[COLOR_COL]).strip()   == str(row2[COLOR_COL]).strip()
#             clarity_match = str(row1[CLARITY_COL]).strip() == str(row2[CLARITY_COL]).strip()

#             table_match  = abs(float(row1[TABLE_COL])  - float(row2[TABLE_COL]))  <= table_tol
#             depth_match  = abs(float(row1[DEPTH_COL])  - float(row2[DEPTH_COL]))  <= depth_tol
#             length_match = abs(float(row1[LENGTH_COL]) - float(row2[LENGTH_COL])) <= length_tol
#             width_match  = abs(float(row1[WIDTH_COL])  - float(row2[WIDTH_COL]))  <= width_tol

#             if (
#                 shape_match and color_match and clarity_match
#                 and table_match and depth_match
#                 and length_match and width_match
#             ):
#                 pid = str(pair_number).zfill(2)
#                 df.at[i, "Pair_ID"] = f"{pid}A"
#                 df.at[j, "Pair_ID"] = f"{pid}B"
#                 used_rows.add(i)
#                 used_rows.add(j)
#                 pair_number += 1
#                 break

#         except Exception:
#             continue

# total_pairs = pair_number - 1
# total_paired = total_pairs * 2
# total_unpaired = len(df) - total_paired


import networkx as nx

df["Pair_ID"] = ""

# ==========================================
# BUILD MATCHING GRAPH
# ==========================================

G = nx.Graph()

for i in range(len(df)):
    G.add_node(i)

for i in range(len(df)):

    row1 = df.iloc[i]

    for j in range(i + 1, len(df)):

        row2 = df.iloc[j]

        try:

            shape_match = (
                str(row1[SHAPE_COL]).strip()
                ==
                str(row2[SHAPE_COL]).strip()
            )

            color_match = (
                str(row1[COLOR_COL]).strip()
                ==
                str(row2[COLOR_COL]).strip()
            )

            clarity_match = (
                str(row1[CLARITY_COL]).strip()
                ==
                str(row2[CLARITY_COL]).strip()
            )

            table_diff = abs(
                float(row1[TABLE_COL])
                -
                float(row2[TABLE_COL])
            )

            depth_diff = abs(
                float(row1[DEPTH_COL])
                -
                float(row2[DEPTH_COL])
            )

            length_diff = abs(
                float(row1[LENGTH_COL])
                -
                float(row2[LENGTH_COL])
            )

            width_diff = abs(
                float(row1[WIDTH_COL])
                -
                float(row2[WIDTH_COL])
            )

            table_match = table_diff <= table_tol
            depth_match = depth_diff <= depth_tol
            length_match = length_diff <= length_tol
            width_match = width_diff <= width_tol

            if (
                shape_match
                and color_match
                and clarity_match
                and table_match
                and depth_match
                and length_match
                and width_match
            ):

                score = (
                    100
                    - table_diff
                    - depth_diff
                    - (length_diff * 10)
                    - (width_diff * 10)
                )

                G.add_edge(
                    i,
                    j,
                    weight=score
                )

        except Exception:
            continue

# ==========================================
# MAXIMUM MATCHING
# ==========================================

matching = nx.max_weight_matching(
    G,
    maxcardinality=True
)

pair_number = 1

for i, j in matching:

    pid = str(pair_number).zfill(2)

    df.at[i, "Pair_ID"] = f"{pid}A"
    df.at[j, "Pair_ID"] = f"{pid}B"

    pair_number += 1

total_pairs = len(matching)
total_paired = total_pairs * 2
total_unpaired = len(df) - total_paired

# ──────────────────────────────────────────
# MOVE Pair_ID TO FIRST COLUMN
# ──────────────────────────────────────────

other_cols = [c for c in df.columns if c != "Pair_ID"]
df = df[["Pair_ID"] + other_cols]

# ──────────────────────────────────────────
# METRICS
# ──────────────────────────────────────────

st.divider()
st.markdown(
    "<div class='section-head'>📊 Results Summary</div>",
    unsafe_allow_html=True,
)

m1, m2, m3, m4 = st.columns(4)
m1.metric("Total Diamonds", len(df))
m2.metric("✅ Pairs Found", total_pairs)
m3.metric("💎 Paired Stones", total_paired)
m4.metric("⬜ Unpaired", total_unpaired)

# ──────────────────────────────────────────
# STYLED TABLE
# ──────────────────────────────────────────

st.divider()
st.markdown(
    "<div class='section-head'>🔍 Pair Result Table</div>",
    unsafe_allow_html=True,
)


def highlight_pairs(row):
    if row["Pair_ID"] != "":
        return ["background-color: #1a3a2a; color: #9ae6b4;"] * len(row)
    return [""] * len(row)


styled_df = df.style.apply(highlight_pairs, axis=1)
st.dataframe(styled_df, use_container_width=True, height=400)

if total_pairs == 0:
    st.warning(
        "⚠️ No pairs found with current tolerances. "
        "Try increasing the tolerance values in the sidebar."
    )
else:
    st.success(f"✅ Found **{total_pairs} pairs** ({total_paired} stones matched).")

# ──────────────────────────────────────────
# EXPORT TO EXCEL
# ──────────────────────────────────────────

def build_excel(dataframe: pd.DataFrame) -> bytes:
    output = BytesIO()

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        dataframe.to_excel(writer, index=False, sheet_name="Diamond Pairs")

        wb  = writer.book
        ws  = writer.sheets["Diamond Pairs"]

        # ── Column widths ──
        for col_idx, col_name in enumerate(dataframe.columns, start=1):
            try:
                col_max = int(dataframe[col_name].astype(str).map(len).max()) if len(dataframe) else 0
            except Exception:
                col_max = 10
            max_len = max(len(str(col_name)), col_max)
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 30)

        # ── Header style ──
        header_fill   = PatternFill(start_color="1a1a2e", end_color="1a1a2e", fill_type="solid")
        header_font   = Font(color="63B3ED", bold=True, size=11)
        header_align  = Alignment(horizontal="center", vertical="center")
        thin_border   = Border(
            bottom=Side(style="thin", color="63B3ED"),
        )

        for cell in ws[1]:
            cell.fill      = header_fill
            cell.font      = header_font
            cell.alignment = header_align
            cell.border    = thin_border

        ws.row_dimensions[1].height = 22

        # ── Fills for paired rows ──
        # ───────────────────────────────
        # UNIQUE COLOR FOR EACH PAIR
        # ───────────────────────────────

        import random

        pair_col_idx = dataframe.columns.get_loc("Pair_ID") + 1

        pair_fill_map = {}

        def generate_bright_color():
            while True:
                r = random.randint(120, 255)
                g = random.randint(120, 255)
                b = random.randint(120, 255)

                color = f"{r:02X}{g:02X}{b:02X}"

                if color not in pair_fill_map.values():
                    return color

        for row_idx in range(2, len(dataframe) + 2):

            pair_value = ws.cell(
                row=row_idx,
                column=pair_col_idx
            ).value

            if not pair_value:
                continue

            pair_number_key = str(pair_value)[:-1]

            if pair_number_key not in pair_fill_map:

                color = generate_bright_color()

                pair_fill_map[pair_number_key] = PatternFill(
                    start_color=color,
                    end_color=color,
                    fill_type="solid"
                )

            fill = pair_fill_map[pair_number_key]

            for col_idx in range(
                1,
                len(dataframe.columns) + 1
            ):
                ws.cell(
                    row=row_idx,
                    column=col_idx
                ).fill = fill

        pair_font = Font(
            color="000000",
            bold=True
        )

        for row_idx in range(2, len(dataframe) + 2):

            pair_value = ws.cell(
                row=row_idx,
                column=pair_col_idx
            ).value

            if pair_value:

                ws.cell(
                    row=row_idx,
                    column=pair_col_idx
                ).font = pair_font

        ws.freeze_panes = "A2"

    output.seek(0)
    return output.read()


st.divider()
st.markdown(
    "<div class='section-head'>📥 Export</div>",
    unsafe_allow_html=True,
)

# Separate paired and unpaired
paired_df   = df[df["Pair_ID"] != ""].copy()
unpaired_df = df[df["Pair_ID"] == ""].copy()

# Sort paired rows by Pair_ID
paired_df["Pair_Sort"] = (
    paired_df["Pair_ID"]
    .str.extract(r"(\d+)")
    .astype(int)
)

paired_df["AB_Sort"] = (
    paired_df["Pair_ID"]
    .str[-1]
    .map({"A": 0, "B": 1})
)

paired_df = paired_df.sort_values(
    ["Pair_Sort", "AB_Sort"]
)

paired_df = paired_df.drop(
    columns=["Pair_Sort", "AB_Sort"]
)

# Final dataframe: paired first, then unpaired
df_export = pd.concat(
    [paired_df, unpaired_df],
    ignore_index=True
)

excel_bytes = build_excel(df_export)

col_dl, col_info = st.columns([2, 5])
with col_dl:
    st.download_button(
        label="📥 Download Paired Excel",
        data=excel_bytes,
        file_name="Diamond_Pairs.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
with col_info:
    st.markdown(
        f"<p style='color:#94a3b8;font-size:0.85rem;margin-top:0.6rem'>"
        f"Pair_ID column is first. Paired rows grouped together with unique colors. "
        f"Header row is frozen for easy scrolling.</p>",
        unsafe_allow_html=True,
    )
